import streamlit as st
import pandas as pd
import plotly.express as px
import time
import ifcopenshell
import numpy as np
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

# Function to process IFC file
def process_ifc(file):
    ifc_file = ifcopenshell.open(file)
    rebars = ifc_file.by_type('IfcReinforcingBar')
    rebar_data = []
    
    for rebar in rebars:
        name = rebar.Name  
        diameter = getattr(rebar, 'NominalDiameter', None)
        length = round(getattr(rebar, 'BarLength', 0), 1) if getattr(rebar, 'BarLength', None) else 'N/A'
        
        rebar_data.append({
            'Rebar Name': name,
            'Nominal Diameter(mm)': diameter,
            'Bar Length(mm)': length
        })
    
    return pd.DataFrame(rebar_data)

# Function to export summary to formatted Excel
def export_to_excel(df):
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    
    for r_idx, row in enumerate(dataframe_to_rows(df, index=True, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Format header row
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    # Auto-adjust column width
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2
    
    wb.save(output)
    output.seek(0)
    return output

# Streamlit App
st.title('Rebar Data Summary & Visualization')

# File uploader
uploaded_file = st.file_uploader('Upload an IFC file', type=['ifc'])

if uploaded_file:
    with open('temp.ifc', 'wb') as f:
        f.write(uploaded_file.read())
    df_rebars = process_ifc('temp.ifc')
    
    # Refine the Data
    df_refined = df_rebars.copy()
    
    # Extract "Shape", "Steel Grade", and "Bend Radius"
    df_refined['Shape'] = df_refined['Rebar Name'].str.extract(r'Shape\s*([A-Z]\d*)')
    df_refined['Steel Grade'] = df_refined['Rebar Name'].str.extract(r'(K\d{3}C-T)')
    df_refined['Bend Radius (mm)'] = df_refined['Rebar Name'].str.extract(r'R=(\d+)').astype(float)
    
    # Calculate Cross Sectional Area (m²)
    df_refined['Cross Section Area (m²)'] = np.pi * (df_refined['Nominal Diameter(mm)'] / 2000) ** 2
    
    # Calculate Estimated Weight (kg)
    steel_density = 7850  # kg/m³
    df_refined['Weight (kg)'] = df_refined['Cross Section Area (m²)'] * (df_refined['Bar Length(mm)'] / 1000) * steel_density
    
    # Reorder columns for clarity
    df_refined = df_refined[
        ['Rebar Name', 'Shape', 'Steel Grade', 'Bend Radius (mm)', 'Nominal Diameter(mm)', 
         'Bar Length(mm)', 'Cross Section Area (m²)', 'Weight (kg)']
    ]
    
    # Generate summary
    summary = df_refined.groupby('Nominal Diameter(mm)').agg(
        Total_Length_m=('Bar Length(mm)', lambda x: x.sum() / 1000),  # Convert mm to meters
        Total_Weight_kg=('Weight (kg)', lambda x: round(x.sum(), 1)),  # Round weight to 1 decimal place
        Count=('Rebar Name', 'count')
    )
    
    # Show summary table
    st.subheader('Summary Table')
    st.write(summary)
    
    # Button to export summary to formatted Excel
    excel_data = export_to_excel(summary)
    st.download_button(
        label="Download Summary as Excel",
        data=excel_data,
        file_name="summary.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    # Create separate bar plots
    fig_length = px.bar(
        summary, 
        x=summary.index, 
        y='Total_Length_m', 
        title='Total Length by Nominal Diameter', 
        labels={'Total_Length_m': 'Total Length (m)', 'Nominal Diameter(mm)': 'Nominal Diameter (mm)'}
    )
    st.plotly_chart(fig_length)
    
    fig_weight = px.bar(
        summary, 
        x=summary.index, 
        y='Total_Weight_kg', 
        title='Total Weight by Nominal Diameter', 
        labels={'Total_Weight_kg': 'Total Weight (kg)', 'Nominal Diameter(mm)': 'Nominal Diameter (mm)'}
    )
    st.plotly_chart(fig_weight)
    
    # Show raw data (optional)
    if st.checkbox('Show raw data'):
        st.write(df_refined)
else:
    st.info("Please upload an IFC file to process and visualize the data.")
